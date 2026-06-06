#!/usr/bin/env python3
"""
舵机扫描数据可视化查看器 (GUI 折线图)
======================================
从 servo_sweep_test.py 生成的 .xlsx 文件读取数据，
显示交互式折线图，支持缩放、平移、数据探查。

用法:
  python view_sweep_data.py                              # 默认读取 py/servo_sweep_result.xlsx
  python view_sweep_data.py -f py/my_sweep.xlsx           # 指定文件
  python view_sweep_data.py -f py/servo_sweep_result.xlsx --raw  # 同时显示原始帧数据

依赖: pip install matplotlib openpyxl
"""

import sys
import argparse
import numpy as np

# ====================== 数据加载 ======================

def load_from_xlsx(filepath: str) -> dict:
    """从 xlsx 读取扫描数据, 返回 {angles, rates, stdevs, speeds, raw_data, fit_info}"""
    try:
        import openpyxl
    except ImportError:
        print("[错误] 需要 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # --- Sheet 1: 扫描汇总 ---
    ws1 = wb["扫描汇总"]
    angles, rates, stdevs, speeds, ox, oy = [], [], [], [], [], []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        # col: 1=采样点, 2=RT_TO, 3=舵机角度, 4=yaw_rate, 5=std, 6=speed, 7=frames, 8=odom_x, 9=odom_y
        if row[2] is None:
            break
        angles.append(float(row[2]))
        rates.append(float(row[3]))
        stdevs.append(float(row[4]))
        speeds.append(float(row[5]))
        ox.append(float(row[7]) if row[7] is not None else 0.0)
        oy.append(float(row[8]) if row[8] is not None else 0.0)

    # --- Sheet 2: 线性拟合 ---
    fit_info = {}
    try:
        ws2 = wb["线性拟合"]
        for row in ws2.iter_rows(min_row=2, max_row=6, values_only=True):
            if row[0] and row[1]:
                fit_info[str(row[0])] = str(row[1])
    except KeyError:
        pass

    # --- Sheet 3: 原始帧数据 ---
    raw_data = {}  # {servo_angle: [(frame_idx, gyro, speed), ...]}
    try:
        ws3 = wb["原始帧数据"]
        for row in ws3.iter_rows(min_row=2, values_only=True):
            if row[2] is None:
                break
            servo = float(row[2])  # col 3
            frame = int(row[3])     # col 4
            gyro = float(row[4])    # col 5
            speed = float(row[5])   # col 6
            if servo not in raw_data:
                raw_data[servo] = []
            raw_data[servo].append((frame, gyro, speed))
    except KeyError:
        pass

    wb.close()
    return {
        "angles": np.array(angles),
        "rates": np.array(rates),
        "stdevs": np.array(stdevs),
        "speeds": np.array(speeds),
        "odom_x": np.array(ox),
        "odom_y": np.array(oy),
        "raw_data": raw_data,
        "fit_info": fit_info,
    }


# ====================== 绘图 ======================

def plot_data(data: dict, show_raw: bool = False, title_suffix: str = ""):
    """启动交互式 matplotlib GUI"""
    import matplotlib
    matplotlib.use("TkAgg")  # 独立 GUI 窗口
    import matplotlib.pyplot as plt
    from matplotlib.widgets import Button, RadioButtons

    # 中文字体
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

    angles = data["angles"]
    rates = data["rates"]
    stdevs = data["stdevs"]
    speeds = data["speeds"]
    ox = data["odom_x"]
    oy = data["odom_y"]
    raw_data = data["raw_data"]
    fit_info = data["fit_info"]

    # 配色
    C_BLUE = "#2196F3"
    C_RED = "#F44336"
    C_GREEN = "#4CAF50"
    C_GRAY = "#9E9E9E"
    C_ORANGE = "#FF9800"

    if show_raw:
        # ====== 3 子图布局: 汇总 + 原始帧 + 速度 ======
        fig = plt.figure("Servo Sweep Data Viewer", figsize=(16, 10))
        gs = fig.add_gridspec(2, 2, hspace=0.3, wspace=0.3)
        ax_main = fig.add_subplot(gs[0, 0])     # 汇总: angle vs yaw_rate
        ax_raw = fig.add_subplot(gs[0, 1])       # 原始帧: time series
        ax_speed = fig.add_subplot(gs[1, 0])     # 速度变化
        ax_odom = fig.add_subplot(gs[1, 1])      # 里程计轨迹
    else:
        # ====== 2 子图布局: 汇总 + 速度 ======
        fig = plt.figure("Servo Sweep Data Viewer", figsize=(14, 8))
        gs = fig.add_gridspec(1, 2, wspace=0.25)
        ax_main = fig.add_subplot(gs[0, 0])
        ax_speed = fig.add_subplot(gs[0, 1])
        ax_raw = ax_odom = None

    fig.suptitle(f"舵机转向线性度扫描 {title_suffix}".strip(),
                 fontsize=14, fontweight="bold", y=0.98)

    # ---- 子图1: angle → yaw_rate (带误差棒) ----
    ax_main.errorbar(angles, rates, yerr=stdevs, fmt='o-', color=C_BLUE,
                     ecolor=C_GRAY, capsize=3, markersize=7, linewidth=1.5,
                     markerfacecolor=C_BLUE, markeredgecolor='white', markeredgewidth=1,
                     label="实测数据")
    ax_main.axvline(x=90, color=C_GRAY, linestyle="--", linewidth=1, alpha=0.6, label="中位 90°")
    ax_main.axhline(y=0, color=C_GRAY, linestyle="--", linewidth=1, alpha=0.6)

    # 线性拟合线
    if fit_info:
        try:
            slope_str = fit_info.get("灵敏度 (斜率)", "")
            intercept_str = fit_info.get("截距", "")
            if slope_str and intercept_str:
                slope = float(slope_str)
                intercept = float(intercept_str)
                x_fit = np.linspace(min(angles), max(angles), 100)
                y_fit = slope * (x_fit - 90) + intercept
                ax_main.plot(x_fit, y_fit, '--', color=C_RED, linewidth=1.5, alpha=0.8,
                            label=f"拟合: {slope:.2f} dps/°")
        except (ValueError, KeyError):
            pass

    ax_main.set_xlabel("舵机角度 (°)", fontsize=11)
    ax_main.set_ylabel("偏航角速度 (dps)", fontsize=11)
    ax_main.set_title("转向特性曲线 (Angle → Yaw Rate)", fontsize=12)
    ax_main.legend(loc="upper left", fontsize=9)
    ax_main.grid(True, alpha=0.3)

    # 标注数据点数值
    for a, r in zip(angles, rates):
        if abs(r) > 1:  # 只标注明显的偏航
            ax_main.annotate(f"{r:.1f}", (a, r), textcoords="offset points",
                           xytext=(0, 10), fontsize=7, color=C_BLUE, alpha=0.7,
                           ha="center")

    # ---- 子图2: 速度 vs 舵机角度 ----
    ax_speed.plot(angles, speeds, 's-', color=C_ORANGE, markersize=7,
                  linewidth=1.5, markerfacecolor=C_ORANGE, markeredgecolor='white')
    ax_speed.set_xlabel("舵机角度 (°)", fontsize=11)
    ax_speed.set_ylabel("平均车速", fontsize=11)
    ax_speed.set_title("车速 vs 舵机角度 (转弯减速效应)", fontsize=12)
    ax_speed.grid(True, alpha=0.3)

    # ---- 子图3: 原始帧时序 (按舵机角度着色) ----
    if ax_raw is not None and raw_data:
        # 只显示部分有代表性的角度 (每隔一个)
        unique_angles = sorted(raw_data.keys())
        cmap = plt.cm.coolwarm
        for i, sa in enumerate(unique_angles):
            if i % max(1, len(unique_angles) // 12) != 0 and sa not in [min(unique_angles), max(unique_angles), 90.0]:
                continue
            frames = raw_data[sa]
            if not frames:
                continue
            idxs = [f[0] for f in frames]
            gyros = [f[1] for f in frames]
            color = cmap(i / max(1, len(unique_angles) - 1))
            ax_raw.plot(idxs, gyros, '-', color=color, linewidth=0.8, alpha=0.8,
                       label=f"{sa:.0f}°")
        ax_raw.set_xlabel("帧序号", fontsize=11)
        ax_raw.set_ylabel("gyro_dps", fontsize=11)
        ax_raw.set_title("各角度原始帧: 角速度时序", fontsize=12)
        ax_raw.legend(loc="upper right", fontsize=6, ncol=2)
        ax_raw.grid(True, alpha=0.3)

    # ---- 子图4: 里程计轨迹 ----
    if ax_odom is not None:
        ax_odom.plot(ox, oy, 'o-', color=C_GREEN, markersize=7, linewidth=1.5,
                    markerfacecolor=C_GREEN, markeredgecolor='white')
        ax_odom.set_xlabel("odom_x (cm)", fontsize=11)
        ax_odom.set_ylabel("odom_y (cm)", fontsize=11)
        ax_odom.set_title("里程计轨迹 (转弯时的 XY 位移)", fontsize=12)
        ax_odom.grid(True, alpha=0.3)
        ax_odom.axhline(y=0, color=C_GRAY, linestyle="--", alpha=0.4)
        ax_odom.axvline(x=0, color=C_GRAY, linestyle="--", alpha=0.4)
        # 标注起点
        if len(ox) > 0:
            ax_odom.annotate("起点", (ox[0], oy[0]), fontsize=9, color=C_GREEN)

    # ---- 信息面板 (用 fig.text 显示拟合参数) ----
    info_lines = []
    if fit_info:
        for k, v in fit_info.items():
            info_lines.append(f"{k}: {v}")
    info_text = "\n".join(info_lines) if info_lines else "无拟合数据"
    fig.text(0.02, 0.01, info_text, fontsize=8, family="monospace",
             bbox=dict(boxstyle="round,pad=0.5", facecolor="lightyellow", alpha=0.8))

    plt.tight_layout(rect=[0, 0.06, 1, 0.95])
    print(f"\n[GUI] 交互窗口已打开 (缩放: 滚轮 | 平移: 拖拽 | 保存: 工具栏磁盘图标)")
    print(f"[GUI] 关闭窗口退出")
    plt.show()


# ====================== 入口 ======================

def main():
    parser = argparse.ArgumentParser(description="舵机扫描数据 GUI 折线图查看器")
    parser.add_argument("--file", "-f", default="py/servo_sweep_result.xlsx",
                        help="xlsx 文件路径 (默认: py/servo_sweep_result.xlsx)")
    parser.add_argument("--raw", "-r", action="store_true",
                        help="显示原始帧时序数据 (4子图模式)")
    parser.add_argument("--title", "-t", default="", help="图表标题后缀")
    args = parser.parse_args()

    print(f"[加载] {args.file}")
    try:
        data = load_from_xlsx(args.file)
    except FileNotFoundError:
        print(f"[错误] 文件不存在: {args.file}")
        print(f"  请先运行 servo_sweep_test.py 生成数据文件")
        sys.exit(1)
    except KeyError as e:
        print(f"[错误] xlsx 格式不匹配: 缺少 Sheet {e}")
        sys.exit(1)

    n_angles = len(data["angles"])
    n_raw = sum(len(v) for v in data["raw_data"].values())
    print(f"[数据] {n_angles} 个采样点, {n_raw} 帧原始数据")

    if data["fit_info"]:
        slope = data["fit_info"].get("灵敏度 (斜率)", "?")
        print(f"[拟合] 灵敏度 = {slope} dps/°")

    plot_data(data, show_raw=args.raw, title_suffix=args.title)


if __name__ == "__main__":
    main()
