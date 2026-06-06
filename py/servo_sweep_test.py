#!/usr/bin/env python3
"""
舵机转向线性插值扫描测试
========================
不修改固件，用串口 RT_TO 命令控制舵机，测量 angle → yaw_rate 映射关系。

流程: SR_ACC加速(不发DT_STA,舵机自由) → RT_TO扫描 → SR_PAU停车
扫描路径: 90°(中) → 120°(左极) → 60°(右极) → 90°(回中)
采样间隔: 每5°一个采样点，插值过渡0.5s，稳态采集1.5s

用法:
  python servo_sweep_test.py                      # 默认 COM7
  python servo_sweep_test.py --port COM8          # 指定串口
  python servo_sweep_test.py --speed 3 --plot     # 3档速度 + 自动画图
"""

import sys
import os
import time
import json
import argparse
import threading
from collections import deque
from dataclasses import dataclass, field
from typing import Optional, List, Tuple
from datetime import datetime

import serial
import serial.tools.list_ports

# ====================== 配置 ======================
DEFAULT_PORT = "COM7"
DEFAULT_BAUDRATE = 9600
DEFAULT_TIMEOUT = 0.1

# RT_TO 命令格式: RT_TOxxx → Angle = 180 - xxx
# 所以 RT_TO090 → servo=90°, RT_TO060 → servo=120°(左), RT_TO120 → servo=60°(右)
RT_TO = lambda s: f"RT_TO{s:03d}"  # s = 180 - servo_angle

# 扫描参数
SWEEP_LEFT_MAX  = 60   # RT_TO值 → 舵机=120° (左转极限)
SWEEP_RIGHT_MAX = 120  # RT_TO值 → 舵机=60°  (右转极限)
SWEEP_CENTER    = 90   # RT_TO值 → 舵机=90°  (中位)
SWEEP_STEP      = 5    # 采样间隔 (°)
HOLD_TIME       = 1.5  # 每采样点稳态采集时间 (秒)
RAMP_FRAMES     = 25   # 插值过渡帧数 (25帧 × 20ms = 0.5s)
RAMP_INTERVAL   = 0.02 # 插值帧间隔 (20ms)

# 数据采集
MIN_SPEED = 0.3  # 最低稳定速度, 低于此值认为车速不够


@dataclass
class SweepSample:
    """单个采样点的汇总数据"""
    rt_to_value: int       # RT_TO命令值
    servo_angle: float     # 实际舵机角度 (180 - rt_to_value)
    avg_yaw_rate: float    # 平均偏航角速度 (dps)
    std_yaw_rate: float    # 角速度标准差
    avg_speed: float       # 平均车速
    avg_odom_x: float      # 平均横向位移
    avg_odom_y: float      # 平均纵向位移
    frame_count: int       # 采集帧数
    raw_frames: list = field(default_factory=list)  # 原始数据帧


class ServoLinearSweep:
    """舵机线性扫描测试器"""

    def __init__(self, port: str = DEFAULT_PORT, baudrate: int = DEFAULT_BAUDRATE):
        self.port = port
        self.baudrate = baudrate
        self.ser: Optional[serial.Serial] = None
        self._rx_buffer = ""
        self._lock = threading.Lock()
        self.results: List[SweepSample] = []

    # ---- 串口 ----
    def open(self) -> bool:
        try:
            self.ser = serial.Serial(self.port, self.baudrate, timeout=DEFAULT_TIMEOUT)
            print(f"[串口] {self.port} 已打开 @ {self.baudrate}bps")
            return True
        except Exception as e:
            print(f"[串口] 打开失败: {e}")
            return False

    def close(self):
        if self.ser and self.ser.is_open:
            self.ser.close()

    def send_cmd(self, cmd: str):
        """发送@指令"""
        if self.ser and self.ser.is_open:
            self.ser.write(f"@{cmd}\r\n".encode())
            # 不打印每个 RT_TO 减少刷屏, 只打关键命令
            if not cmd.startswith("RT_TO"):
                print(f"  [TX] {cmd}")

    def read_frame(self) -> Optional[dict]:
        """读取一帧数据, 返回 dict{gyro, yaw, servo, speed, odom_x, odom_y} 或 None"""
        if not self.ser or not self.ser.is_open:
            return None
        try:
            with self._lock:
                waiting = self.ser.in_waiting
                if waiting > 0:
                    chunk = self.ser.read(waiting).decode("utf-8", errors="ignore")
                    self._rx_buffer += chunk
                if "\n" in self._rx_buffer:
                    line, self._rx_buffer = self._rx_buffer.split("\n", 1)
                    parts = line.strip().split(",")
                    if len(parts) >= 6:
                        return {
                            "gyro_dps": float(parts[0]),
                            "yaw_deg": float(parts[1]),
                            "servo_deg": float(parts[2]),
                            "speed": float(parts[3]),
                            "odom_x": float(parts[4]),
                            "odom_y": float(parts[5]),
                        }
        except (ValueError, IndexError):
            pass
        return None

    # ---- 小车控制 ----
    def start_moving(self, speed_acc: int = 2):
        """加速到稳定速度 (不发DT_STA, 舵机保持在90°, 不进入直行PID模式)"""
        # 确保舵机先回中 (不发DT_STA, 直接设置角度)
        print("[控制] 舵机回中...")
        self.send_cmd("RT_TO090")
        time.sleep(0.3)

        print(f"[控制] 加速 × {speed_acc}...")
        for i in range(speed_acc):
            self.send_cmd("SR_ACC")
            time.sleep(0.3)

        # 等待速度稳定
        print("[控制] 等待车速稳定...")
        for _ in range(50):
            f = self.read_frame()
            if f and f["speed"] > MIN_SPEED:
                print(f"  车速={f['speed']:.1f}, 开始扫描")
                return True
            time.sleep(0.02)
        print("  [警告] 未检测到速度, 继续执行")
        return False

    def stop_moving(self):
        """停车"""
        print("[控制] 停车...")
        self.send_cmd("SR_PAU")
        time.sleep(0.2)
        self.send_cmd("RT_TO090")  # 舵机回中
        time.sleep(0.2)

    # ---- 扫描核心 ----
    def ramp_to(self, from_rt: int, to_rt: int):
        """线性插值从 from_rt 过渡到 to_rt, 共 RAMP_FRAMES 帧"""
        for i in range(RAMP_FRAMES):
            t = (i + 1) / RAMP_FRAMES  # 0.04, 0.08, ..., 1.0
            rt_val = int(from_rt + (to_rt - from_rt) * t)
            self.send_cmd(RT_TO(rt_val))
            time.sleep(RAMP_INTERVAL)

    def collect_at(self, rt_val: int, hold_time: float = HOLD_TIME) -> SweepSample:
        """在某个 RT_TO 值稳态采集数据"""
        # 先跳到目标角度
        self.send_cmd(RT_TO(rt_val))
        time.sleep(0.3)  # 等舵机到位 + 车体响应

        frames = []
        deadline = time.time() + hold_time
        while time.time() < deadline:
            f = self.read_frame()
            if f:
                frames.append(f)
            time.sleep(0.01)  # 避免空转

        # 汇总统计
        if frames:
            gyro_vals = [f["gyro_dps"] for f in frames]
            speed_vals = [f["speed"] for f in frames]
            odom_x_vals = [f["odom_x"] for f in frames]
            odom_y_vals = [f["odom_y"] for f in frames]
            n = len(gyro_vals)

            avg_gyro = sum(gyro_vals) / n
            avg_speed = sum(speed_vals) / n
            avg_ox = sum(odom_x_vals) / n
            avg_oy = sum(odom_y_vals) / n
            std_gyro = (sum((g - avg_gyro) ** 2 for g in gyro_vals) / n) ** 0.5
        else:
            avg_gyro = std_gyro = avg_speed = avg_ox = avg_oy = 0.0
            n = 0

        return SweepSample(
            rt_to_value=rt_val,
            servo_angle=180.0 - rt_val,
            avg_yaw_rate=avg_gyro,
            std_yaw_rate=std_gyro,
            avg_speed=avg_speed,
            avg_odom_x=avg_ox,
            avg_odom_y=avg_oy,
            frame_count=n,
            raw_frames=[{"gyro": f["gyro_dps"], "speed": f["speed"]} for f in frames],
        )

    def run_sweep(self, speed_acc: int = 2, step_deg: int = SWEEP_STEP, hold_time: float = HOLD_TIME, output_path: str = "py/servo_sweep_result.xlsx"):
        """执行完整扫描流程"""
        print(f"\n{'='*60}")
        print(f"舵机转向线性插值扫描")
        print(f"{'='*60}")
        print(f"  扫描范围: {SWEEP_LEFT_MAX}(左极) ← {SWEEP_CENTER}(中) → {SWEEP_RIGHT_MAX}(右极)")
        print(f"  采样步长: {SWEEP_STEP}°")
        print(f"  插值帧数: {RAMP_FRAMES} ({RAMP_FRAMES * RAMP_INTERVAL:.1f}s)")
        print(f"  稳态采集: {HOLD_TIME}s/点")
        print(f"{'='*60}\n")

        # 构建扫描路径: 中 → 左极 → 右极 → 中
        left_sweep  = list(range(SWEEP_CENTER, SWEEP_LEFT_MAX - 1, -step_deg))
        right_sweep = list(range(SWEEP_LEFT_MAX + step_deg, SWEEP_RIGHT_MAX + 1, step_deg))
        back_center = list(range(SWEEP_RIGHT_MAX - step_deg, SWEEP_CENTER - 1, -step_deg))
        full_path = left_sweep + right_sweep + back_center

        # 去重连续重复值
        waypoints = []
        for v in full_path:
            if not waypoints or v != waypoints[-1]:
                waypoints.append(v)

        print(f"扫描路径 ({len(waypoints)} 个采样点):")
        rt_labels = [f"{w:03d}(舵机{180-w}°)" for w in waypoints]
        for i in range(0, len(rt_labels), 6):
            print(f"  {' → '.join(rt_labels[i:i+6])}")
        print()

        # 启动小车
        if not self.start_moving(speed_acc):
            self.stop_moving()
            return

        # 执行扫描
        last_rt = SWEEP_CENTER
        total = len(waypoints)
        try:
            for idx, rt_val in enumerate(waypoints):
                servo_deg = 180 - rt_val
                direction = "左" if servo_deg > 90 else ("右" if servo_deg < 90 else "中")
                print(f"\n[{idx+1}/{total}] 目标: RT_TO{rt_val:03d} → 舵机={servo_deg}° ({direction}转)")

                # 插值过渡
                if rt_val != last_rt:
                    print(f"  插值过渡 {last_rt:03d}→{rt_val:03d} ({RAMP_FRAMES}帧)...")
                    self.ramp_to(last_rt, rt_val)

                # 稳态采集
                print(f"  稳态采集 {HOLD_TIME}s...")
                sample = self.collect_at(rt_val, hold_time)
                self.results.append(sample)

                print(f"  → yaw_rate={sample.avg_yaw_rate:+.2f}±{sample.std_yaw_rate:.2f} dps, "
                      f"speed={sample.avg_speed:.2f}, frames={sample.frame_count}")

                last_rt = rt_val

                # 检查是否走偏太远 (安全保护)
                if abs(sample.avg_yaw_rate) > 200:
                    print("  [警告] 角速度异常大, 可能已经失控, 中断扫描")
                    break

        except KeyboardInterrupt:
            print("\n[扫描] 用户中断")
        finally:
            self.stop_moving()

        # 输出结果
        self.print_summary()
        self.save_xlsx(filepath=output_path)

    def print_summary(self):
        """输出扫描结果"""
        if not self.results:
            return

        print(f"\n{'='*60}")
        print(f"扫描结果 ({len(self.results)} 个采样点)")
        print(f"{'='*60}")
        print(f"{'RT_TO':>6s} {'舵机°':>7s} {'yaw_rate':>10s} {'±std':>8s} {'speed':>8s} {'frames':>7s}")
        print("-" * 52)

        for s in self.results:
            print(f"{s.rt_to_value:6d} {s.servo_angle:7.1f} {s.avg_yaw_rate:10.2f} "
                  f"{s.std_yaw_rate:8.2f} {s.avg_speed:8.2f} {s.frame_count:7d}")

        # 提取线性区
        print(f"\n--- 线性区分析 ---")
        # 筛选 |yaw_rate| > 0.5 且不是极值点 (排除死区和饱和区)
        valid = [(s.servo_angle, s.avg_yaw_rate) for s in self.results
                 if abs(s.avg_yaw_rate) > 0.5 and 65 < s.servo_angle < 115]
        if valid:
            angles, rates = zip(*valid)
            print(f"  有效采样点: {len(valid)}")
            print(f"  角度范围: {min(angles):.0f}° ~ {max(angles):.0f}°")
            print(f"  角速度范围: {min(rates):.1f} ~ {max(rates):.1f} dps")

            # 简单线性拟合
            n = len(angles)
            sum_x = sum(angles)
            sum_y = sum(rates)
            sum_xy = sum(a * r for a, r in valid)
            sum_x2 = sum(a * a for a in angles)
            slope = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x * sum_x)
            intercept = (sum_y - slope * sum_x) / n
            print(f"  拟合: yaw_rate = {slope:.3f} × (angle - 90) + {intercept:.2f}")
            print(f"  灵敏度: {slope:.2f} dps/°")

        # JSON 输出
        print(f"\n--- JSON 数据 ---")
        json_out = [
            {
                "rt_to": s.rt_to_value,
                "servo_deg": round(s.servo_angle, 1),
                "yaw_rate_dps": round(s.avg_yaw_rate, 3),
                "yaw_std": round(s.std_yaw_rate, 3),
                "speed": round(s.avg_speed, 3),
                "frames": s.frame_count,
            }
            for s in self.results
        ]
        print(json.dumps(json_out, indent=2, ensure_ascii=False))

    def save_xlsx(self, filepath: str = "servo_sweep_result.xlsx"):
        """将扫描结果保存为 .xlsx 文件 (需要 openpyxl 库)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.chart import ScatterChart, Reference, Series
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("[xlsx] 需要 openpyxl 库: pip install openpyxl")
            return

        if not self.results:
            print("[xlsx] 无数据可保存")
            return

        wb = openpyxl.Workbook()

        # ---- 样式 ----
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center")

        def style_header(ws, row, cols):
            for c in range(1, cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align

        def style_data(ws, start_row, end_row, cols):
            for r in range(start_row, end_row + 1):
                for c in range(1, cols + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = thin_border
                    cell.alignment = center_align

        # ====== Sheet 1: 汇总 ======
        ws1 = wb.active
        ws1.title = "扫描汇总"
        headers1 = ["采样点", "RT_TO值", "舵机角度(°)", "平均偏航角速度(dps)",
                     "角速度标准差", "平均车速", "采集帧数", "平均里程X(cm)", "平均里程Y(cm)"]
        for c, h in enumerate(headers1, 1):
            ws1.cell(row=1, column=c, value=h)
        style_header(ws1, 1, len(headers1))

        for i, s in enumerate(self.results):
            row = i + 2
            ws1.cell(row=row, column=1, value=i + 1)
            ws1.cell(row=row, column=2, value=s.rt_to_value)
            ws1.cell(row=row, column=3, value=round(s.servo_angle, 1))
            ws1.cell(row=row, column=4, value=round(s.avg_yaw_rate, 3))
            ws1.cell(row=row, column=5, value=round(s.std_yaw_rate, 3))
            ws1.cell(row=row, column=6, value=round(s.avg_speed, 3))
            ws1.cell(row=row, column=7, value=s.frame_count)
            ws1.cell(row=row, column=8, value=round(s.avg_odom_x, 3))
            ws1.cell(row=row, column=9, value=round(s.avg_odom_y, 3))
        style_data(ws1, 2, len(self.results) + 1, len(headers1))

        # 自动列宽
        for c in range(1, len(headers1) + 1):
            ws1.column_dimensions[get_column_letter(c)].width = 20

        # 散点图: 舵机角度 vs 偏航角速度
        chart = ScatterChart()
        chart.title = "舵机角度 vs 偏航角速度"
        chart.x_axis.title = "舵机角度 (°)"
        chart.y_axis.title = "偏航角速度 (dps)"
        chart.style = 10
        x_values = Reference(ws1, min_col=3, min_row=2, max_row=len(self.results) + 1)
        y_values = Reference(ws1, min_col=4, min_row=2, max_row=len(self.results) + 1)
        series = Series(y_values, x_values, title="yaw_rate")
        series.marker.symbol = "circle"
        series.marker.size = 6
        chart.series.append(series)
        ws1.add_chart(chart, "K2")

        # ====== Sheet 2: 线性拟合 ======
        ws2 = wb.create_sheet("线性拟合")
        valid = [(s.servo_angle, s.avg_yaw_rate, s.std_yaw_rate) for s in self.results
                 if abs(s.avg_yaw_rate) > 0.5 and 65 < s.servo_angle < 115]
        if valid:
            angles, rates, _ = zip(*valid)
            n = len(angles)
            sum_x = sum(angles)
            sum_y = sum(rates)
            sum_xy = sum(a * r for a, r in zip(angles, rates))
            sum_x2 = sum(a * a for a in angles)
            slope = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x * sum_x)
            intercept = (sum_y - slope * sum_x) / n

            fit_headers = ["参数", "值", "单位"]
            for c, h in enumerate(fit_headers, 1):
                ws2.cell(row=1, column=c, value=h)
            style_header(ws2, 1, 3)

            fit_data = [
                ("线性区角度范围", f"{min(angles):.1f} ~ {max(angles):.1f}", "°"),
                ("灵敏度 (斜率)", f"{slope:.4f}", "dps/°"),
                ("截距", f"{intercept:.4f}", "dps"),
                ("线性公式", f"yaw_rate = {slope:.3f} × (angle - 90) + {intercept:.2f}", ""),
                ("有效采样点数", f"{n}", "个"),
            ]
            for i, (label, val, unit) in enumerate(fit_data):
                row = i + 2
                ws2.cell(row=row, column=1, value=label)
                ws2.cell(row=row, column=2, value=val)
                ws2.cell(row=row, column=3, value=unit)

            # 拟合曲线散点图
            chart2 = ScatterChart()
            chart2.title = "线性拟合: 舵机 vs 偏航角速度"
            chart2.x_axis.title = "舵机角度 (°)"
            chart2.y_axis.title = "偏航角速度 (dps)"
            chart2.style = 10

            # 写原始数据用于画图
            ws2.cell(row=10, column=1, value="原始数据(线性区)")
            ws2.cell(row=10, column=2, value="拟合值")
            ws2.cell(row=11, column=1, value="舵机角度")
            ws2.cell(row=11, column=2, value="yaw_rate")
            ws2.cell(row=11, column=3, value="拟合yaw_rate")
            for i, (a, r, _) in enumerate(valid):
                row = i + 12
                ws2.cell(row=row, column=1, value=round(a, 1))
                ws2.cell(row=row, column=2, value=round(r, 3))
                ws2.cell(row=row, column=3, value=round(slope * (a - 90) + intercept, 3))

            n_valid = len(valid)
            x_ref = Reference(ws2, min_col=1, min_row=12, max_row=11 + n_valid)
            y_ref = Reference(ws2, min_col=2, min_row=12, max_row=11 + n_valid)
            fit_ref = Reference(ws2, min_col=3, min_row=12, max_row=11 + n_valid)

            raw_series = Series(y_ref, x_ref, title="实测")
            raw_series.marker.symbol = "circle"
            raw_series.marker.size = 6
            chart2.series.append(raw_series)

            fit_series = Series(fit_ref, x_ref, title="拟合")
            fit_series.marker.symbol = "none"
            chart2.series.append(fit_series)

            ws2.add_chart(chart2, "E2")

            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 20
            ws2.column_dimensions['C'].width = 18

        # ====== Sheet 3: 原始帧数据 ======
        ws3 = wb.create_sheet("原始帧数据")
        raw_headers = ["采样点", "RT_TO", "舵机角度", "帧序号", "gyro_dps", "speed"]
        for c, h in enumerate(raw_headers, 1):
            ws3.cell(row=1, column=c, value=h)
        style_header(ws3, 1, len(raw_headers))

        row = 2
        for i, s in enumerate(self.results):
            for j, f in enumerate(s.raw_frames):
                ws3.cell(row=row, column=1, value=i + 1)
                ws3.cell(row=row, column=2, value=s.rt_to_value)
                ws3.cell(row=row, column=3, value=round(s.servo_angle, 1))
                ws3.cell(row=row, column=4, value=j + 1)
                ws3.cell(row=row, column=5, value=round(f["gyro"], 3))
                ws3.cell(row=row, column=6, value=round(f["speed"], 3))
                row += 1
        style_data(ws3, 2, row - 1, len(raw_headers))
        for c in range(1, len(raw_headers) + 1):
            ws3.column_dimensions[get_column_letter(c)].width = 14

        # 保存
        wb.save(filepath)
        print(f"[xlsx] 已保存 {filepath}")
        print(f"        Sheet 1: 扫描汇总 ({len(self.results)} 个采样点 + 散点图)")
        print(f"        Sheet 2: 线性拟合")
        print(f"        Sheet 3: 原始帧数据 ({row - 2} 帧)")


# ====================== 入口 ======================
def main():
    parser = argparse.ArgumentParser(description="舵机转向线性插值扫描测试")
    parser.add_argument("--port", "-p", default=DEFAULT_PORT, help=f"串口号 (默认: {DEFAULT_PORT})")
    parser.add_argument("--baudrate", "-b", type=int, default=DEFAULT_BAUDRATE, help=f"波特率")
    parser.add_argument("--speed", "-s", type=int, default=2, help="加速档数 (默认: 2)")
    parser.add_argument("--step", type=int, default=SWEEP_STEP, help=f"采样步长° (默认: {SWEEP_STEP})")
    parser.add_argument("--hold", type=float, default=HOLD_TIME, help=f"每点稳态采集秒数 (默认: {HOLD_TIME})")
    parser.add_argument("--plot", action="store_true", help="扫描完成后画图")
    parser.add_argument("--output", "-o", default="py/servo_sweep_result.xlsx",
                        help="xlsx 输出文件路径 (默认: servo_sweep_result.xlsx)")
    args = parser.parse_args()

    sweeper = ServoLinearSweep(port=args.port, baudrate=args.baudrate)
    if not sweeper.open():
        sys.exit(1)

    try:
        sweeper.run_sweep(speed_acc=args.speed, step_deg=args.step, hold_time=args.hold, output_path=args.output)

        if args.plot and sweeper.results:
            try:
                import matplotlib.pyplot as plt
                angles = [s.servo_angle for s in sweeper.results]
                rates = [s.avg_yaw_rate for s in sweeper.results]
                stdevs = [s.std_yaw_rate for s in sweeper.results]

                fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

                ax1.errorbar(angles, rates, yerr=stdevs, fmt='o-', capsize=4, markersize=6)
                ax1.axvline(x=90, color='gray', linestyle='--', alpha=0.5, label='中位90°')
                ax1.axhline(y=0, color='gray', linestyle='--', alpha=0.5)
                ax1.set_xlabel('Servo Angle (deg)')
                ax1.set_ylabel('Yaw Rate (dps)')
                ax1.set_title('Servo Angle vs Yaw Rate (转向线性度)')
                ax1.grid(True, alpha=0.3)
                ax1.legend()

                ax2.plot(angles, rates, 'o-', markersize=6)
                ax2.set_xlabel('Servo Angle (deg)')
                ax2.set_ylabel('Yaw Rate (dps)')
                ax2.set_title('Yaw Rate Detail')
                ax2.grid(True, alpha=0.3)

                plt.tight_layout()
                plt.savefig('servo_sweep_result.png', dpi=150)
                print(f"\n[绘图] 已保存 servo_sweep_result.png")
                plt.show()
            except ImportError:
                print("[绘图] 需要 matplotlib: pip install matplotlib")

    finally:
        sweeper.close()


if __name__ == "__main__":
    main()
